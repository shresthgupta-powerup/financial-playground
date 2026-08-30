"""Financial Planning Playground — Streamlit host for the handed-off engine.

v1 ground rule: the engine under code/app/planning/ is a byte-identical copy of
the CRM handoff (see 00_START_HERE.md); this file is ONLY a UI wrapper. The form
model (defaults, goal templates, progressive disclosure, risk-profile mapping)
is a direct port of code/frontend/planForm.js, and the output shaping mirrors
the pure helpers in code/app/planning/service.py (_build_snapshot,
_build_goal_results, _build_wealth_monthly, append_csv_summary_cols) — service
itself is not imported because it pulls the app-coupled DB modules.
"""

import copy
import json
import re
import sys
import uuid
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent / "code"))

import pandas as pd
import streamlit as st

from app.planning import (
    ENGINE_SOURCE_SHA,
    GLIDEPATH_VERSION,
    PlanValidationError,
    find_retirement_date,
    get_glide_paths,
    run_simulation,
    validate_plan_config,
)
import app.planning as _planning

# Fallback covers a hot-reloaded app running against a stale cached engine
# module (Streamlit Cloud pulls code without restarting Python until reboot).
ENGINE_UPDATED = getattr(_planning, "ENGINE_UPDATED", "2026-08-24")
from app.planning.advisor_export import build_advisor_workbook
from app.planning.engine import (
    _DEFAULT_INSTRUMENT_PARAMS,
    _resolve_goals,
    _resolve_recurring_occurrences,
    expand_recurring_goal_to_tranches,
    format_inr,
)
from app.planning.goal_grid import goal_sleeves
from app.planning.schemas import RISK_PROFILE_CORE_RETURNS

# ── Picklists (mirror planForm.js) ──────────────────────────────────────────
GOAL_TYPES = ["Non-Negotiable", "Semi-Negotiable", "Negotiable"]

# CRM goals contract (Punit, 2026-08-30). The CRM stores one FLAT row per
# goal; tokens are exact and case-sensitive - anything else rejects the goal.
CRM_GOAL_CATEGORIES = [
    None, "education", "marriage", "home", "vehicle", "travel",
    "retirement", "healthcare", "emergency", "business", "other",
]
CRM_NEGOTIABILITY = {
    "Non-Negotiable": "non_negotiable",
    "Semi-Negotiable": "semi_negotiable",
    "Negotiable": "negotiable",
}
CRM_FREQUENCY = {
    "Monthly": "monthly", "Quarterly": "quarterly",
    "Half-Yearly": "half_yearly", "Annual": "yearly",
}
# Their marker for a series whose length we cannot state (our Lifetime goals).
CRM_OPEN_ENDED_OCCURRENCES = 500
GOAL_NATURES = ["Non-replenishing", "Replenishing"]

# v2 (2026-08-24): "nature" is no longer asked. A goal's purpose is DERIVED —
# does it still have cashflows beyond the grid's reach when this plan ends?
# The form asks Structure (one-time vs recurring) and Type (negotiability);
# GOAL_NATURES / NATURE_FROM_DISPLAY survive only so saved files from every
# earlier era still load (their stored nature is read and ignored).
NATURE_FROM_DISPLAY = {"Lumpsum": "Non-replenishing", "Recurring": "Replenishing"}
GOAL_STRUCTURES = ["Lumpsum", "Recurring"]
GOAL_START_MODES = ["Fixed", "At retirement"]
GOAL_END_MODES = ["Occurrences", "Fixed date", "Lifetime"]
INVESTMENT_END_MODES = ["At retirement", "Fixed"]
RECURRING_FREQUENCIES = ["Monthly", "Quarterly", "Half-Yearly", "Annual"]
STEPUP_FREQUENCIES = ["Annual", "Half-Yearly", "Quarterly", "Monthly"]
RISK_PROFILES = list(RISK_PROFILE_CORE_RETURNS.keys())

MONTH_NAMES = [
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December",
]

GOAL_TEMPLATES = {
    "Retirement Income": "retirement_income",
    "Child Education": "child_education",
    "Marriage": "marriage",
    "Home Purchase": "home_purchase",
    "Custom": "custom",
}


# ── Date helpers (month-grid invariant: every date is the 1st of a month) ───
def month_start_today() -> pd.Timestamp:
    now = pd.Timestamp.today()
    return pd.Timestamp(now.year, now.month, 1)


def add_years(ts: pd.Timestamp, years: int) -> pd.Timestamp:
    return pd.Timestamp(ts.year + years, ts.month, 1)


def fmt_mon_yyyy(ts) -> str:
    return pd.Timestamp(ts).strftime("%b %Y")


def short_inr(amount) -> str:
    v = float(amount)
    if v >= 1e7:
        return f"₹{v / 1e7:.2f} Cr"
    if v >= 1e5:
        return f"₹{v / 1e5:.2f} L"
    return format_inr(v)


def inr_hint(amount) -> str:
    """Grouped INR + lakh/crore hint (planForm UX note 6)."""
    try:
        v = float(amount)
    except (TypeError, ValueError):
        return ""
    if v >= 1e7:
        return f"{format_inr(v)}  ({v / 1e7:.2f} Cr)"
    if v >= 1e5:
        return f"{format_inr(v)}  ({v / 1e5:.2f} L)"
    return format_inr(v)


# ── Default form state (mirror planForm.makeDefault*) ───────────────────────
def make_default_stream(index: int, today: pd.Timestamp) -> dict:
    return {
        "name": "Salary" if index == 0 else f"Stream {index + 1}",
        "amount": 100_000 if index == 0 else 50_000,
        "start_date": today,
        "end_date_mode": "At retirement",
        "end_date": add_years(today, 30 if index == 0 else 20),
        "step_up_percent": 10.0,
        "step_up_frequency": "Annual",
        "step_up_date": today,
    }


def make_default_goal(index: int, today: pd.Timestamp) -> dict:
    return {
        "name": f"Goal {index + 1}",
        "description": "",
        "type": "Non-Negotiable",
        "nature": "Non-replenishing",
        "structure": "Lumpsum",
        "start_date_mode": "Fixed",
        "start_date": add_years(today, 15),
        "amount": 1_000_000,
        "frequency": "Annual",
        "end_mode": "Occurrences",
        "occurrences": 1,
        "end_date": None,
        "inflation_percent": 6.0,
        # CRM contract fields. purpose_id is minted by the CRM (None until a
        # goal has been uploaded once) and must ride back on every re-export.
        "purpose_id": None,
        "goal_category": None,
    }


def make_goal_from_template(template_key: str, index: int, today: pd.Timestamp) -> dict:
    base = make_default_goal(index, today)
    if template_key == "retirement_income":
        base.update(
            name="Retirement Income", description="Monthly income post-retirement",
            nature="Replenishing", structure="Recurring",
            start_date_mode="At retirement", start_date=add_years(today, 30),
            amount=75_000, frequency="Monthly", end_mode="Lifetime",
            occurrences=360, end_date=None, inflation_percent=6.0,
            goal_category="retirement",
        )
    elif template_key == "child_education":
        base.update(
            name="Child Education", description="Annual education fees",
            nature="Non-replenishing", structure="Recurring", type="Non-Negotiable",
            start_date_mode="Fixed", start_date=add_years(today, 12),
            amount=1_500_000, frequency="Annual", end_mode="Occurrences",
            occurrences=4, inflation_percent=8.0, goal_category="education",
        )
    elif template_key == "marriage":
        base.update(
            name="Marriage", description="Wedding expenses",
            nature="Non-replenishing", structure="Lumpsum", type="Semi-Negotiable",
            start_date_mode="Fixed", start_date=add_years(today, 20),
            amount=3_000_000, inflation_percent=7.0, goal_category="marriage",
        )
    elif template_key == "home_purchase":
        base.update(
            name="Home Purchase", description="Down payment / purchase",
            nature="Non-replenishing", structure="Lumpsum", type="Negotiable",
            start_date_mode="Fixed", start_date=add_years(today, 8),
            amount=5_000_000, inflation_percent=6.0, goal_category="home",
        )
    return base


def unique_goal_name(name: str, existing) -> str:
    """First use keeps the name; later duplicates get ' 2', ' 3', ...

    Goal names identify a goal's funding chain in the engine, so two goals
    sharing a name used to collide (fixed engine-side by +goaldedupe, which
    labels the second '<name> #2'). We make the names distinct HERE, in the
    form, so the CM sees and controls them instead of discovering a rename in
    the Excel later.
    """
    name = (name or "Goal").strip() or "Goal"
    existing = set(existing)
    if name not in existing:
        return name
    n = 2
    while f"{name} {n}" in existing:
        n += 1
    return f"{name} {n}"


def duplicate_goal_names(goals) -> list:
    """Names used by more than one goal (order preserved)."""
    seen, dupes = set(), []
    for g in goals:
        nm = (g.get("name") or "").strip()
        if nm and nm in seen and nm not in dupes:
            dupes.append(nm)
        seen.add(nm)
    return dupes


def goal_purpose(goal: dict, current_date) -> str:
    """Derived purpose (§4.5): does this goal outlive the plan's carve window?

    Never an input — it falls out of the cashflow series and the grid's reach
    for the goal's negotiability (5 / 4 / 3 years). Read it as a statement
    about TODAY: a one-time goal further out than its column's reach is
    "beyond the window" right now and enters it in a later month, exactly as a
    40-year income stream's later occurrences do. That is why the wording is
    about the window and not about the goal being endless.
    """
    # Resolve the series length the way the engine does, so the derivation
    # sees the real cashflows even BEFORE a run: a Lifetime series (or one
    # starting at a not-yet-known retirement) outlives every carve window by
    # construction, and a Fixed-date end implies its occurrence count.
    # Without this, unresolved goals derived from a 1-occurrence stub -
    # Retirement Income exported as "Non-replenishing" in the as-entered JSON.
    if goal.get("structure") == "Recurring":
        if (goal.get("end_mode") == "Lifetime"
                or goal.get("start_date_mode") == "At retirement"):
            return "Extends beyond the funding window"
    try:
        g = dict(goal)
        if (g.get("structure") == "Recurring"
                and g.get("end_mode") == "Fixed date"
                and g.get("end_date") is not None):
            g["occurrences"] = _resolve_recurring_occurrences(g, None)
        g["negotiability"] = g.get("type")
        g["structure"] = "recurring" if g.get("structure") == "Recurring" else "one-time"
        g["occurrences"] = g.get("occurrences") or 1
        g["frequency"] = (g.get("frequency") or "monthly").lower()
        g["inflation"] = float(g.get("inflation_percent", 0) or 0) / 100.0
        sleeves = goal_sleeves(g, current_date, apply_tax=False)
    except Exception:
        return ""
    return ("Extends beyond the funding window"
            if sleeves["purpose"] == "GOAL_REPLENISH"
            else "Fully inside the funding window")


def payments_fixed_for(goal: dict) -> bool:
    """POLICY, not preference (operator decision, 2026-08-25 - no CM option).

    Every recurring goal is contract-fixed - the amount escalates at the
    growth % only until the FIRST payment, then all payments stay at that
    amount (fees lock at admission, EMIs are signed) - EXCEPT income-like
    series, which must keep tracking cost of living. "Income" is structural,
    not a name: the goal starts at retirement, or its payments run for life.
    """
    if goal.get("structure") != "Recurring":
        return False
    is_income = (goal.get("end_mode") == "Lifetime"
                 or goal.get("start_date_mode") == "At retirement")
    return not is_income


def normalise_goal(goal: dict) -> dict:
    """Progressive-disclosure reset (mirror planForm.normaliseGoal).

    v2: Structure is the goal's own input — a stored ``nature`` from an older
    saved file no longer overrides it.
    """
    g = dict(goal)
    if g.get("structure") not in GOAL_STRUCTURES:
        g["structure"] = "Lumpsum"
    if g["structure"] == "Lumpsum":
        g["frequency"] = None
        g["end_mode"] = None
        g["occurrences"] = 1
        g["end_date"] = None
    else:
        if g.get("frequency") not in RECURRING_FREQUENCIES:
            g["frequency"] = "Monthly"
        if g.get("end_mode") not in GOAL_END_MODES:
            g["end_mode"] = "Occurrences"
        if g["end_mode"] == "Occurrences":
            g["end_date"] = None
            if not g.get("occurrences") or g["occurrences"] < 1:
                g["occurrences"] = 1
        elif g["end_mode"] == "Fixed date":
            if not g.get("occurrences"):
                g["occurrences"] = 1
        else:  # Lifetime
            g["end_date"] = None
    return g


# ── Output shaping (ports of service.py's pure helpers) ─────────────────────
def resolve_instrument_params(risk_profile: str) -> dict:
    merged = copy.deepcopy(_DEFAULT_INSTRUMENT_PARAMS)
    merged["core_corpus"]["return"] = RISK_PROFILE_CORE_RETURNS.get(
        risk_profile, RISK_PROFILE_CORE_RETURNS["Balanced"]
    )
    return merged


def build_snapshot(comprehensive_df: pd.DataFrame, retirement_date: pd.Timestamp) -> dict | None:
    df = comprehensive_df.copy()
    df["Date"] = pd.to_datetime(df["Date"])
    snap = df[df["Date"] >= retirement_date].head(1)
    if snap.empty:
        return None
    row = snap.iloc[0]
    core = float(row.get("Core Corpus Value", 0) or 0)
    debt = float(row.get("Debt Pool Value", 0) or 0)
    hybrid = float(row.get("Hybrid Pool Value", 0) or 0)
    # v2: "Debt Pool Value" is ALL debt held for goals, and the per-goal
    # "<goal> Debt Value" columns are its breakdown - the same rupees seen two
    # ways. Adding both double-counts every rupee of goal money (it did not in
    # v1, where the pools and the per-goal chains were genuinely separate
    # money). Total = the three aggregates; the per-goal sums are reported
    # alongside as the split, never added in.
    goal_debt = float(sum(row.get(c, 0) or 0 for c in row.index if c.endswith(" Debt Value")))
    goal_hybrid = float(sum(row.get(c, 0) or 0 for c in row.index if c.endswith(" Hybrid Value")))
    total = core + debt + hybrid
    return {
        "core": round(core, 2), "debt": round(debt, 2), "hybrid": round(hybrid, 2),
        "goal_debt": round(goal_debt, 2), "goal_hybrid": round(goal_hybrid, 2),
        "total": round(total, 2),
    }


def build_goal_results(config: dict, retirement_date) -> pd.DataFrame:
    current_date = pd.Timestamp(config["current_date"])
    rows = []
    for goal in config.get("goals", []) or []:
        if goal.get("start_date_mode", "Fixed").lower() == "at retirement" and retirement_date is not None:
            start = pd.Timestamp(retirement_date)
        elif goal.get("start_date") is not None:
            start = pd.Timestamp(goal["start_date"])
        else:
            start = current_date
        years = max(0.0, (start - current_date).days / 365.25)
        pv = float(goal.get("amount", 0) or 0)
        inflation = float(goal.get("inflation_percent", 0) or 0)
        rows.append({
            "Goal": goal.get("name", ""),
            "Purpose (derived)": goal_purpose(goal, current_date),
            "Structure": goal.get("structure", ""),
            "Starts": fmt_mon_yyyy(start),
            "Amount (today's ₹)": format_inr(pv),
            "Amount at start (FV)": format_inr(pv * ((1 + inflation / 100) ** years)),
        })
    return pd.DataFrame(rows)


def wealth_frame(comprehensive_df: pd.DataFrame, death_date: pd.Timestamp) -> pd.DataFrame:
    df = comprehensive_df.copy()
    df["Date"] = pd.to_datetime(df["Date"])
    df = df[df["Date"] <= death_date]
    # Only the three aggregate columns; the per-goal columns break the SAME
    # money down by goal and would double-count (see build_snapshot).
    out = pd.DataFrame({"Date": df["Date"]})
    out["Total wealth"] = sum(
        df[c].fillna(0) for c in _POOL_VALUE_COLS if c in df.columns)
    out["Core corpus"] = df.get("Core Corpus Value", 0)
    out["Debt pool"] = df.get("Debt Pool Value", 0)
    out["Hybrid pool"] = df.get("Hybrid Pool Value", 0)
    return out.set_index("Date")


_POOL_VALUE_COLS = {"Core Corpus Value", "Debt Pool Value", "Hybrid Pool Value"}


def csv_with_summary(comprehensive_df: pd.DataFrame) -> bytes:
    """Port of service.append_csv_summary_cols + to_csv."""
    out = comprehensive_df.copy()
    value_cols = [c for c in out.columns if c.endswith("Value")]
    goal_value_cols = [c for c in value_cols if c not in _POOL_VALUE_COLS]
    # Aggregates only - the per-goal columns are a breakdown of the same money.
    out["Total Wealth (Rs)"] = sum(
        out[c].fillna(0) for c in _POOL_VALUE_COLS if c in out.columns)
    out["Goal Tranches (Rs)"] = out[goal_value_cols].fillna(0).sum(axis=1) if goal_value_cols else 0.0
    return out.to_csv(index=False).encode("utf-8")


def new_simulation_id() -> str:
    """Readable, sortable, unique: SIM-<IST timestamp>-<random suffix>."""
    stamp = pd.Timestamp.now(tz="Asia/Kolkata").strftime("%Y%m%d-%H%M%S")
    return f"SIM-{stamp}-{uuid.uuid4().hex[:6]}"


def inputs_filename(config: dict) -> str:
    """Unique, human-sortable filename: client + timestamp, filesystem-safe."""
    raw = (config.get("client_name") or "plan").strip() or "plan"
    slug = re.sub(r"[^A-Za-z0-9]+", "-", raw).strip("-") or "plan"
    stamp = pd.Timestamp.now().strftime("%Y-%m-%d_%H%M")
    return f"financial_plan_inputs_{slug}_{stamp}.json"


def _iso_or_none(d):
    """Serialise a date to 'YYYY-MM-01' (month grid), or None."""
    if d is None:
        return None
    try:
        return pd.Timestamp(d).strftime("%Y-%m-01")
    except Exception:
        return None


# ── CRM goal-import contract (CRM team, 2026-07-29) ─────────────────────────
# The goals array of this file is imported into the CRM as official goals, so
# it follows the CRM's strict field contract:
# - nature is Replenishing / Non-replenishing (the same vocabulary the UI
#   shows again since the 2026-08-14 revert).
# - structure accompanies Non-replenishing goals only (ignored for
#   Replenishing); end_mode is never exported on resolved goals — a concrete
#   occurrences count takes its place.
# - type / frequency casing standardised to the CRM picklists.
_TYPE_TO_CRM = {"Non-Negotiable": "Non-negotiable", "Semi-Negotiable": "Semi-negotiable"}
_FREQ_TO_CRM = {"Half-Yearly": "Half-yearly"}
_FREQ_MONTHS = {"Monthly": 1, "Quarterly": 3, "Half-Yearly": 6, "Annual": 12}


def _crm_goal(g: dict, resolved: bool, current_date=None) -> dict:
    """One goals-array entry, keys ordered as in the CRM's example.

    resolved=True (successful run): contract shape — start_date_mode is always
    "Fixed" with the actual date ("At retirement" resolved to the solved
    retirement date by engine._resolve_goals), occurrences is the concrete
    count ("Lifetime" collapsed against the plan's death date), and end_date is
    the computed last-payment date. resolved=False (infeasible run — nothing
    solved to resolve against): the goal is recorded exactly as entered.
    """
    goal = {"name": g.get("name")}
    if g.get("description"):
        goal["description"] = g["description"]
    goal["type"] = _TYPE_TO_CRM.get(g.get("type"), g.get("type"))
    # §4.4: nature is not an input — export the DERIVED value so the CRM
    # contract keeps its field without anyone typing it.
    _derived = goal_purpose(g, pd.Timestamp(current_date) if current_date is not None
                            else pd.Timestamp.today())
    goal["nature"] = ("Replenishing" if _derived.startswith("Extends")
                      else "Non-replenishing")
    goal["structure"] = g.get("structure")
    recurring = g.get("structure") == "Recurring"
    freq = g.get("frequency")
    if resolved:
        start = pd.Timestamp(g["start_date"])
        goal["start_date_mode"] = "Fixed"
        goal["start_date"] = start.strftime("%Y-%m-01")
        goal["amount"] = g.get("amount")
        if recurring:
            occ = int(g.get("occurrences") or 1)
            goal["frequency"] = _FREQ_TO_CRM.get(freq, freq)
            goal["occurrences"] = occ
            months = _FREQ_MONTHS.get(freq)
            if months:
                goal["end_date"] = (
                    start + pd.DateOffset(months=months * (occ - 1))
                ).strftime("%Y-%m-01")
    else:
        goal["start_date_mode"] = g.get("start_date_mode")
        goal["start_date"] = _iso_or_none(g.get("start_date"))
        goal["amount"] = g.get("amount")
        if recurring:
            goal["frequency"] = _FREQ_TO_CRM.get(freq, freq)
            goal["occurrences"] = g.get("occurrences")
            goal["end_mode"] = g.get("end_mode")
            goal["end_date"] = _iso_or_none(g.get("end_date"))
    goal["inflation_percent"] = g.get("inflation_percent")
    if recurring:
        goal["payments_fixed_at_start"] = bool(g.get("payments_fixed_at_start"))
    return goal


def crm_goals_upload_json(config: dict, retirement_date) -> bytes:
    """The CRM's goals upload file (Punit's spec, 2026-08-30).

    Strictly ``{"goals": [...]}`` - no envelope - and every goal carries
    exactly the eleven contract keys, always present. Nothing is interpreted
    on their side: a missing key, an unknown key, a null where one is not
    allowed, or a token in the wrong case rejects that goal.

    Requires a solved retirement_date: the contract has no "at retirement"
    concept, so those starts must already be concrete dates.
    """
    current_date = pd.Timestamp(config["current_date"])
    death_date = current_date + pd.DateOffset(
        years=int(float(config.get("target_lifetime", 90))
                  - float(config.get("current_age", 30))))
    resolved = _resolve_goals(config.get("goals", []) or [],
                              pd.Timestamp(retirement_date), death_date)

    goals = []
    for entered, g in zip(config.get("goals", []) or [], resolved):
        recurring = g.get("structure") == "Recurring"
        # Their marker for an unbounded series: our Lifetime end mode. A goal
        # with a real count keeps it, even when it starts at retirement.
        open_ended = recurring and entered.get("end_mode") == "Lifetime"
        if open_ended:
            occurrences = CRM_OPEN_ENDED_OCCURRENCES
        elif recurring:
            occurrences = max(1, int(g.get("occurrences") or 1))
        else:
            occurrences = 1
        goals.append({
            "purpose_id": g.get("purpose_id") or None,
            "goal_name": g.get("name"),
            "goal_type": g.get("goal_category") or None,
            "goal_negotiability": CRM_NEGOTIABILITY.get(g.get("type")),
            "goal_description": g.get("description") or "",
            "amount_per_occurrence": int(round(float(g.get("amount") or 0))),
            "occurrences": occurrences,
            "frequency": (CRM_FREQUENCY.get(g.get("frequency"))
                          if occurrences > 1 else None),
            "start_date": pd.Timestamp(g["start_date"]).strftime("%Y-%m-01"),
            "inflation": round(float(g.get("inflation_percent") or 0) / 100.0, 6),
            "goal_status": "active",
        })
    return json.dumps({"goals": goals}, indent=2).encode("utf-8")


_CRM_NEGOTIABILITY_BACK = {v: k for k, v in CRM_NEGOTIABILITY.items()}
_CRM_FREQUENCY_BACK = {v: k for k, v in CRM_FREQUENCY.items()}


def _goal_from_crm_row(row: dict) -> dict:
    """One flat CRM goal row -> our goal shape (their spec, 2026-08-30).

    ``occurrences == CRM_OPEN_ENDED_OCCURRENCES`` is their marker for a series
    whose length is not stated, which is our Lifetime end mode. Restoring it
    matters: it is what makes ``payments_fixed_for`` classify the goal as
    income again, so a retirement income keeps escalating after a round trip
    through the CRM. Everything else lands as a plain fixed-count series.
    """
    occ = int(row.get("occurrences") or 1)
    open_ended = occ == CRM_OPEN_ENDED_OCCURRENCES
    recurring = occ > 1
    return {
        "name": row.get("goal_name"),
        "description": row.get("goal_description") or "",
        "type": _CRM_NEGOTIABILITY_BACK.get(row.get("goal_negotiability")),
        "structure": "Recurring" if recurring else "Lumpsum",
        "start_date_mode": "Fixed",
        "start_date": row.get("start_date"),
        "amount": row.get("amount_per_occurrence"),
        "frequency": _CRM_FREQUENCY_BACK.get(row.get("frequency")),
        "occurrences": occ,
        "end_mode": ("Lifetime" if open_ended
                     else ("Occurrences" if recurring else None)),
        "end_date": None,
        "inflation_percent": (float(row["inflation"]) * 100.0
                              if row.get("inflation") is not None else None),
        "purpose_id": row.get("purpose_id") or None,
        "goal_category": row.get("goal_type") or None,
    }


def crm_upload_filename(config: dict) -> str:
    raw = (config.get("client_name") or "plan").strip() or "plan"
    slug = re.sub(r"[^A-Za-z0-9]+", "-", raw).strip("-") or "plan"
    return f"crm_goals_upload_{slug}_{pd.Timestamp.now().strftime('%Y-%m-%d_%H%M')}.json"


def build_inputs_json(config: dict, retirement_date=None) -> bytes:
    """Serialise the user's inputs as pretty JSON bytes; goals are CRM-ready.

    The goals array follows the CRM import contract (see _crm_goal): engine
    nature vocabulary, standardised casings and — when retirement_date is given
    (a successful run) — concrete dates and occurrence counts resolved by the
    engine's own _resolve_goals, so the file matches the simulation exactly.
    Everything else stays a faithful record of the form (the CRM reads only
    goals + personal.client_name + generated_at and ignores the rest):
    - dates become 'YYYY-MM-01' strings (the month grid);
    - the internal 'm3_id' the app injects for the engine is dropped.
    """
    goals_in = config.get("goals", []) or []
    if retirement_date is not None:
        death_date = pd.Timestamp(config["current_date"]) + pd.DateOffset(
            years=int(float(config.get("target_lifetime", 90))
                      - float(config.get("current_age", 30)))
        )
        goals_out = [
            _crm_goal(g, resolved=True, current_date=config["current_date"])
            for g in _resolve_goals(goals_in, pd.Timestamp(retirement_date), death_date)
        ]
    else:
        goals_out = [_crm_goal(g, resolved=False,
                               current_date=config["current_date"]) for g in goals_in]
    doc = {
        "generated_at": pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S"),
        "simulation_id": config.get("simulation_id", ""),
        "retirement_mode": config.get("retirement_mode", "earliest"),
        # target_date is authoritative; target_age is derived, for readability.
        "target_date": _iso_or_none(config.get("target_date"))
        if config.get("retirement_mode") == "target_age" else None,
        "target_age": round(target_age_of(config, target_retirement_date(config)), 1)
        if config.get("retirement_mode") == "target_age" and config.get("target_date")
        else None,
        "engine_version": ENGINE_SOURCE_SHA,
        "glidepath_version": GLIDEPATH_VERSION,
        "personal": {
            "client_name": config.get("client_name", ""),
            "phone": config.get("phone", ""),
            "current_date": _iso_or_none(config.get("current_date")),
            "current_age": config.get("current_age"),
            "target_lifetime": config.get("target_lifetime"),
            "current_corpus": config.get("current_corpus"),
            "risk_profile": config.get("risk_profile"),
        },
        "investment_streams": [
            {
                "name": s.get("name"),
                "amount": s.get("amount"),
                "start_date": _iso_or_none(s.get("start_date")),
                "end_date_mode": s.get("end_date_mode"),
                "end_date": _iso_or_none(s.get("end_date")),
                "step_up_percent": s.get("step_up_percent"),
                "step_up_frequency": s.get("step_up_frequency"),
                "step_up_date": _iso_or_none(s.get("step_up_date")),
            }
            for s in config.get("investment_streams", []) or []
        ],
        "goals": goals_out,
        "one_time_investments": [
            {
                "name": w.get("name"),
                "date": _iso_or_none(w.get("date")),
                "amount": w.get("amount"),
            }
            for w in config.get("one_time_investments", []) or []
        ],
    }
    return json.dumps(doc, indent=2, ensure_ascii=False).encode("utf-8")


# ── Load a downloaded inputs JSON back into the form ────────────────────────
_TYPE_FROM_CRM = {v: k for k, v in _TYPE_TO_CRM.items()}
_FREQ_FROM_CRM = {v: k for k, v in _FREQ_TO_CRM.items()}


def form_state_from_inputs(doc: dict):
    """Map an inputs JSON (either export shape) back to form state.

    Tolerant on vocabulary: nature accepts the engine words
    (Replenishing/Non-replenishing) and the display words (Recurring/Lumpsum);
    type/frequency accept both our casing and the CRM casing. Resolved exports
    carry no end_mode — their concrete occurrences count maps to
    end_mode="Occurrences", which reproduces the same payout schedule.
    Returns (personal, streams, goals, one_time); raises ValueError on files
    that aren't an inputs JSON.
    """
    if not isinstance(doc, dict) or "goals" not in doc:
        raise ValueError("not an inputs JSON (missing 'goals')")
    today = month_start_today()
    # A CRM goals file is flat rows with no envelope. Translate it into our
    # shape first so the rest of this function is unchanged - this is how
    # CRM-minted purpose_ids come back to us (their spec, 2026-08-30).
    if doc.get("goals") and isinstance(doc["goals"][0], dict)             and "goal_name" in doc["goals"][0]:
        doc = dict(doc)
        doc.setdefault("personal", {})
        doc["goals"] = [_goal_from_crm_row(r) for r in doc["goals"]]
    if "personal" not in doc:
        raise ValueError("not an inputs JSON (missing 'personal')")

    def ts(v, fallback=None):
        if not v:
            return fallback
        try:
            t = pd.Timestamp(v)
            return pd.Timestamp(t.year, t.month, 1)
        except (ValueError, TypeError):
            return fallback

    def num(v, default):
        """Numeric with an explicit None/empty check — NEVER `or`, which would
        turn a legitimate stored 0 into the default (the inflation-0 bug)."""
        if v is None or v == "":
            return float(default)
        return float(v)

    p = doc.get("personal") or {}
    risk = p.get("risk_profile")
    mode = doc.get("retirement_mode")
    current_date = ts(p.get("current_date"), today)
    current_age = int(num(p.get("current_age"), 30))
    # target_date preferred; age-era files (pre 2026-08-26) carry target_age.
    target_date = ts(doc.get("target_date"))
    if target_date is None and doc.get("target_age") is not None:
        years = int(round(float(doc["target_age"]) - current_age))
        target_date = add_years(current_date, max(years, 1))
    personal = {
        "client_name": str(p.get("client_name") or ""),
        "current_date": current_date,
        "current_age": current_age,
        "target_lifetime": int(num(p.get("target_lifetime"), 90)),
        "current_corpus": int(num(p.get("current_corpus"), 0)),
        "risk_profile": risk if risk in RISK_PROFILES else "Balanced",
        "retirement_mode": mode if mode in ("earliest", "target_age") else "earliest",
        "target_date": target_date or add_years(current_date, 25),
        "phone": normalize_phone(str(p.get("phone") or "")) or "",
    }

    streams = []
    for s in doc.get("investment_streams") or []:
        start = ts(s.get("start_date"), today)
        mode = s.get("end_date_mode")
        streams.append({
            "name": str(s.get("name") or f"Stream {len(streams) + 1}"),
            "amount": int(float(s.get("amount") or 0)),
            "start_date": start,
            "end_date_mode": mode if mode in INVESTMENT_END_MODES else "At retirement",
            "end_date": ts(s.get("end_date"), add_years(start, 20)),
            "step_up_percent": num(s.get("step_up_percent"), 0.0),
            "step_up_frequency": s.get("step_up_frequency")
            if s.get("step_up_frequency") in STEPUP_FREQUENCIES else "Annual",
            "step_up_date": ts(s.get("step_up_date"), start),
        })

    goals = []
    for g in doc.get("goals") or []:
        nature = NATURE_FROM_DISPLAY.get(g.get("nature"), g.get("nature"))
        if nature not in GOAL_NATURES:
            nature = "Non-replenishing"
        gtype = _TYPE_FROM_CRM.get(g.get("type"), g.get("type"))
        if gtype not in GOAL_TYPES:
            gtype = "Non-Negotiable"
        freq = _FREQ_FROM_CRM.get(g.get("frequency"), g.get("frequency"))
        end_mode = g.get("end_mode")
        if end_mode not in GOAL_END_MODES:
            end_mode = "Occurrences"
        mode = g.get("start_date_mode")
        # Structure: the file's value wins for Non-replenishing goals; files
        # from the rename era (no structure key, or display-vocab nature) fall
        # back on Lumpsum unless the goal is recurring by its fields.
        structure = g.get("structure")
        if structure not in GOAL_STRUCTURES:
            structure = "Recurring" if (g.get("frequency") or (g.get("occurrences") or 1) > 1) else "Lumpsum"
        goals.append(normalise_goal({
            "name": str(g.get("name") or f"Goal {len(goals) + 1}"),
            "description": str(g.get("description") or ""),
            "type": gtype,
            "nature": nature,
            "structure": structure,
            "start_date_mode": mode if mode in GOAL_START_MODES else "Fixed",
            "start_date": ts(g.get("start_date"), add_years(today, 15)),
            "amount": int(float(g.get("amount") or 0)),
            "frequency": freq if freq in RECURRING_FREQUENCIES else "Annual",
            "occurrences": int(float(g.get("occurrences") or 1)),
            "end_mode": end_mode,
            "end_date": ts(g.get("end_date")),
            "inflation_percent": num(g.get("inflation_percent"), 6.0),
            "purpose_id": g.get("purpose_id") or None,
            "goal_category": g.get("goal_category") or None,
        }))

    # Saved plans from before the duplicate-name fix can carry two goals with
    # the same name; number them on load so the CM sees them as separate.
    _names = []
    for g in goals:
        g["name"] = unique_goal_name(g.get("name"), _names)
        _names.append(g["name"])

    one_time = []
    for w in doc.get("one_time_investments") or []:
        one_time.append({
            "name": str(w.get("name") or f"Investment {len(one_time) + 1}"),
            "date": ts(w.get("date"), today),
            "amount": int(float(w.get("amount") or 0)),
        })
    return personal, streams, goals, one_time


def _load_doc_into_form(doc: dict, source: str) -> None:
    """Shared by the JSON uploader and the version picker: replace form state.

    Personal widgets are set DIRECTLY via their session-state keys — the
    documented way to change a keyed widget programmatically. The widgets
    themselves carry no defaults, so nothing competes with these values.
    """
    personal, streams, goals, one_time = form_state_from_inputs(doc)
    if not personal.get("phone"):
        # Files from before phone was stored: keep whatever is typed now.
        personal["phone"] = st.session_state.get("p_phone", "")
    for item in streams + goals + one_time:
        item["_uid"] = _next_uid()
    st.session_state.streams = streams
    st.session_state.goals = goals
    st.session_state.one_time = one_time
    st.session_state.personal_defaults = personal
    st.session_state["p_client"] = personal["client_name"]
    cd = personal["current_date"]
    st.session_state["p_curdate_m"] = MONTH_NAMES[cd.month - 1]
    st.session_state["p_curdate_y"] = int(cd.year)
    st.session_state["p_age"] = int(personal["current_age"])
    st.session_state["p_life"] = int(personal["target_lifetime"])
    st.session_state["p_corpus"] = int(personal["current_corpus"])
    st.session_state["p_risk"] = personal["risk_profile"]
    st.session_state["p_mode"] = ("Earliest possible"
                                  if personal["retirement_mode"] == "earliest"
                                  else "At a chosen date")
    td = personal["target_date"]
    st.session_state["p_target_m"] = MONTH_NAMES[td.month - 1]
    st.session_state["p_target_y"] = int(td.year)
    st.session_state["p_phone"] = personal.get("phone", "")
    st.session_state.run_output = None
    st.session_state.upload_msg = (
        "success",
        f"Loaded {source} for “{personal['client_name'] or 'plan'}” — review and press Run simulation.",
    )


def _apply_uploaded_inputs():
    """st.button callback: load the uploaded JSON into the form."""
    file = st.session_state.get("inputs_uploader")
    if file is None:
        st.session_state.upload_msg = ("error", "Choose a JSON file first.")
        return
    try:
        doc = json.loads(file.getvalue().decode("utf-8"))
        _load_doc_into_form(doc, "inputs")
    except (ValueError, UnicodeDecodeError) as e:
        st.session_state.upload_msg = ("error", f"Could not read that file: {e}")


# ── Summary sheet (CM one-glance view) ──────────────────────────────────────
# Inserted as the FIRST sheet of the engine's finished workbook — the engine's
# own export code stays untouched. Any error here degrades to the original
# workbook rather than breaking the download.

def _goal_outflow_rows(config: dict, resolve_date, death_date) -> tuple[list, float]:
    """Per-goal totals using the engine's own tranche expansion (engine-exact)."""
    resolved = _resolve_goals(config.get("goals") or [], resolve_date, death_date)
    rows, grand = [], 0.0
    for g in resolved:
        tranches = expand_recurring_goal_to_tranches(g, config["current_date"])
        total = float(sum(fv for _d, fv in tranches))
        first = min((d for d, _fv in tranches), default=g.get("start_date"))
        fv_first = next((fv for d, fv in tranches if d == first), 0.0)
        grand += total
        rows.append({
            "name": g.get("name", ""), "type": g.get("type", ""),
            "nature": goal_purpose(g, config["current_date"]), "structure": g.get("structure", ""),
            "starts": fmt_mon_yyyy(first) if first is not None else "—",
            "payments": len(tranches),
            "pv": float(g.get("amount", 0) or 0),
            "fv_first": fv_first,
            "total_fv": total,
        })
    return rows, grand


def add_summary_sheet(xlsx_bytes: bytes, config: dict, *, kind: str,
                      retirement_date=None, age_at_retirement=None, failure=None,
                      comp_df=None, snapshot=None, death_date=None,
                      mode_info=None) -> bytes:
    """Prepend a styled Summary sheet to the advisor workbook."""
    try:
        import io
        from openpyxl import load_workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        NAVY, GOLD = "16294B", "C9A227"
        GREEN_BG, GREEN_TX = "E4F2E8", "1E6B3A"
        RED_BG, RED_TX = "FAE6E4", "9C2B23"
        GREY_BG = "F2F5F9"
        thin = Side(style="thin", color="C9D1DC")
        box = Border(left=thin, right=thin, top=thin, bottom=thin)

        wb = load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb.create_sheet("Summary", 0)
        widths = {"A": 30, "B": 17, "C": 15, "D": 15, "E": 13, "F": 11,
                  "G": 15, "H": 16, "I": 16}
        for col, w in widths.items():
            ws.column_dimensions[col].width = w

        def put(cell, value, *, bold=False, size=11, color="1C2430", fill=None,
                align="left", italic=False, border=False):
            c = ws[cell]
            c.value = value
            c.font = Font(name="Calibri", bold=bold, size=size, color=color,
                          italic=italic)
            if fill:
                c.fill = PatternFill("solid", start_color=fill)
            c.alignment = Alignment(horizontal=align, vertical="center",
                                    wrap_text=True)
            if border:
                c.border = box
            return c

        # Title + identity
        ws.merge_cells("A1:I1")
        put("A1", "FINANCIAL PLAN — SUMMARY", bold=True, size=16,
            color="FFFFFF", fill=NAVY)
        ws.row_dimensions[1].height = 26
        ws.merge_cells("A2:I2")
        ident = [config.get("client_name") or "—"]
        if config.get("phone"):
            ident.append(f"phone {config['phone']}")
        ident += [
            f"run {pd.Timestamp.now(tz='Asia/Kolkata').strftime('%d %b %Y, %H:%M IST')}",
            f"risk profile {config.get('risk_profile', 'Balanced')} "
            f"({RISK_PROFILE_CORE_RETURNS.get(config.get('risk_profile'), 0.12) * 100:g}% core)",
            f"engine {ENGINE_SOURCE_SHA}",
        ]
        if config.get("simulation_id"):
            ident.append(config["simulation_id"])
        put("A2", "   ·   ".join(ident), italic=True, size=9, color="5A6472")

        # Verdict banner
        ws.merge_cells("A4:I4")
        ws.row_dimensions[4].height = 24
        mi = mode_info or {}
        if kind == "success" and mi.get("target_age") is not None:
            extra = (f"  ·  earliest possible: {fmt_mon_yyyy(mi['earliest'])}"
                     if mi.get("earliest") is not None else "")
            put("A4", f"FEASIBLE at chosen date {fmt_mon_yyyy(retirement_date)} "
                      f"(age {mi['target_age']:.1f}){extra}",
                bold=True, size=13, color=GREEN_TX, fill=GREEN_BG)
        elif kind == "success":
            put("A4", f"FEASIBLE — earliest retirement {fmt_mon_yyyy(retirement_date)}"
                      f" (age {age_at_retirement:.1f})",
                bold=True, size=13, color=GREEN_TX, fill=GREEN_BG)
        elif kind == "target_infeasible":
            sip = mi.get("sip_needed")
            if sip:
                now = mi.get("monthly_now", 0)
                msg = (f"NOT FUNDABLE at {fmt_mon_yyyy(mi['chosen'])} "
                       f"(age {mi['target_age']:.1f}) — additional SIP needed: "
                       f"{format_inr(sip)}/month (raise {format_inr(now)} → "
                       f"{format_inr(now + sip)})")
            else:
                msg = (f"NOT FUNDABLE at {fmt_mon_yyyy(mi['chosen'])} "
                       f"(age {mi['target_age']:.1f}) — not achievable by extra "
                       f"SIP alone; revisit goal amounts or dates")
            if mi.get("earliest") is not None:
                msg += f"  ·  current plan supports {fmt_mon_yyyy(mi['earliest'])}"
            put("A4", msg, bold=True, size=12, color=RED_TX, fill=RED_BG)
        else:
            f = failure or {}
            when = fmt_mon_yyyy(f["date"]) if f.get("date") else "—"
            put("A4", f"NOT FUNDABLE within the plan horizon — first failure "
                      f"{when}: {f.get('description', 'corpus depletion')} "
                      f"(diagnostic run at latest possible retirement)",
                bold=True, size=12, color=RED_TX, fill=RED_BG)

        # Key numbers
        put("A6", "KEY NUMBERS", bold=True, color=NAVY, fill=GREY_BG)
        ws.merge_cells("A6:I6")
        current_date = pd.Timestamp(config["current_date"])
        monthly_inv = sum(float(s.get("amount", 0) or 0)
                          for s in config.get("investment_streams") or []
                          if pd.Timestamp(s["start_date"]) <= current_date)
        pairs = [("Current corpus", format_inr(config.get("current_corpus", 0))),
                 ("Monthly investment today", format_inr(monthly_inv))]
        totals = None
        if comp_df is not None and not comp_df.empty:
            df = comp_df.copy()
            df["Date"] = pd.to_datetime(df["Date"])
            vc = [c for c in df.columns if c.endswith("Value")]
            totals = df[vc].fillna(0).sum(axis=1)
            peak_i = totals.idxmax()
            if kind == "success":
                years = (pd.Timestamp(retirement_date) - current_date).days / 365.25
                pairs.append(("Years to retirement", f"{years:.1f}"))
                if snapshot:
                    pairs.append(("Wealth at retirement", format_inr(snapshot["total"])))
            pairs += [
                ("Wealth at plan end", format_inr(float(totals.iloc[-1]))),
                ("Peak wealth", f"{format_inr(float(totals.loc[peak_i]))}"
                                f"  ({df['Date'].loc[peak_i].strftime('%b %Y')})"),
            ]
            if "Investment" in df.columns:
                upto = (df["Date"] <= pd.Timestamp(retirement_date)) \
                    if kind == "success" else pd.Series(True, index=df.index)
                pairs.append(("Total investments (to retirement)"
                              if kind == "success" else "Total investments (horizon)",
                              format_inr(float(df.loc[upto, "Investment"].fillna(0).sum()))))
        one_time_total = sum(float(w.get("amount", 0) or 0)
                             for w in config.get("one_time_investments") or [])
        if one_time_total:
            pairs.append(("One-time investments", format_inr(one_time_total)))
        r = 7
        for i, (label, value) in enumerate(pairs):
            col = "A" if i % 2 == 0 else "F"
            vcol = "B" if i % 2 == 0 else "G"
            put(f"{col}{r}", label, size=10, color="5A6472")
            put(f"{vcol}{r}", value, bold=True, size=10)
            if i % 2 == 1:
                r += 1
        if len(pairs) % 2 == 1:
            r += 1

        # Snapshot split (success only)
        if kind == "success" and snapshot:
            r += 1
            put(f"A{r}", "WEALTH AT RETIREMENT — SPLIT", bold=True, color=NAVY,
                fill=GREY_BG)
            ws.merge_cells(f"A{r}:I{r}")
            r += 1
            # goal_debt / goal_hybrid are the per-goal breakdown of the debt
            # and hybrid rows above, not extra money - so they are labelled as
            # a split and excluded from Total.
            for label, key in [("Core corpus", "core"), ("Debt held for goals", "debt"),
                               ("Hybrid held for goals", "hybrid"),
                               ("  of which, by goal (debt)", "goal_debt"),
                               ("  of which, by goal (hybrid)", "goal_hybrid"),
                               ("Total", "total")]:
                put(f"A{r}", label, size=10,
                    color="5A6472" if key != "total" else "1C2430",
                    bold=(key == "total"))
                put(f"B{r}", format_inr(snapshot[key]), bold=(key == "total"), size=10)
                r += 1

        # Goals table
        r += 1
        put(f"A{r}", "GOALS", bold=True, color=NAVY, fill=GREY_BG)
        ws.merge_cells(f"A{r}:I{r}")
        r += 1
        if kind == "success":
            resolve_date = pd.Timestamp(retirement_date)
        elif kind == "target_infeasible" and (mode_info or {}).get("chosen") is not None:
            resolve_date = pd.Timestamp(mode_info["chosen"])
        else:
            resolve_date = pd.Timestamp(death_date)
        goal_rows, grand = _goal_outflow_rows(config, resolve_date, death_date)
        headers = ["Goal", "Type", "Nature", "Structure", "Starts", "Payments",
                   "Amount (today's ₹)", "Cost at start (FV)", "Total outflow (FV)"]
        for j, h in enumerate(headers):
            put(f"{chr(65 + j)}{r}", h, bold=True, size=9, color="FFFFFF",
                fill=NAVY, border=True)
        r += 1
        for g in goal_rows:
            vals = [g["name"], g["type"], g["nature"], g["structure"], g["starts"],
                    g["payments"], format_inr(g["pv"]), format_inr(g["fv_first"]),
                    format_inr(g["total_fv"])]
            for j, v in enumerate(vals):
                put(f"{chr(65 + j)}{r}", v, size=9, border=True,
                    align="right" if j >= 5 else "left")
            r += 1
        put(f"A{r}", "TOTAL", bold=True, size=9, border=True)
        for j in range(1, 8):
            put(f"{chr(65 + j)}{r}", "", border=True)
        put(f"I{r}", format_inr(grand), bold=True, size=9, border=True, align="right")
        r += 2

        # Money in vs money out
        put(f"A{r}", "MONEY IN vs MONEY OUT (full horizon, future value)",
            bold=True, color=NAVY, fill=GREY_BG)
        ws.merge_cells(f"A{r}:I{r}")
        r += 1
        if comp_df is not None and not comp_df.empty and "Investment" in comp_df.columns:
            inv_total = float(comp_df["Investment"].fillna(0).sum())
            put(f"A{r}", "Total investment inflows", size=10, color="5A6472")
            put(f"B{r}", format_inr(inv_total + one_time_total), bold=True, size=10)
            r += 1
        put(f"A{r}", "Total goal outflows", size=10, color="5A6472")
        put(f"B{r}", format_inr(grand), bold=True, size=10)
        r += 2
        put(f"A{r}", "Amounts are engine-computed: goal costs grown from today's "
                     "value at each goal's growth %; every figure matches the "
                     "detailed sheets that follow.",
            italic=True, size=8, color="5A6472")
        ws.merge_cells(f"A{r}:I{r}")

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
    except Exception:
        return xlsx_bytes  # never break the download over a summary problem


# ── Version history (Google Sheet) ──────────────────────────────────────────
# One spreadsheet row per simulation run, keyed by the client's phone number.
# The stored payload is the same inputs JSON the download/upload feature uses,
# so loading a version reuses the exact same restore path. The feature is
# DORMANT until Streamlit secrets provide [gcp_service_account] and
# versions_sheet_id — without them the UI hides and runs save nothing.
# sim_id/output_json were added 2026-08-25 and sit at the END on purpose:
# the header row auto-migrates in place, and appending columns keeps every
# pre-existing row aligned (older rows simply have the two cells empty).
VERSIONS_HEADER = ["phone", "client_name", "saved_at", "version", "note",
                   "result", "inputs_json", "sim_id", "output_json"]


def normalize_phone(raw: str) -> str | None:
    """Digits only; accept +91/0 prefixes by keeping the last 10 digits."""
    digits = re.sub(r"\D", "", raw or "")
    if len(digits) > 10:
        digits = digits[-10:]
    return digits if len(digits) == 10 else None


@st.cache_resource(show_spinner=False)
def _cached_ws():
    try:
        creds = dict(st.secrets["gcp_service_account"])
        sheet_id = st.secrets["versions_sheet_id"]
    except (KeyError, FileNotFoundError):
        return None
    try:
        import gspread
        ws = gspread.service_account_from_dict(creds).open_by_key(sheet_id).sheet1
        if ws.row_values(1) != VERSIONS_HEADER:
            ws.update(values=[VERSIONS_HEADER], range_name="A1")
        return ws
    except Exception as e:
        st.session_state.versions_error = str(e)
        return None


def _versions_ws():
    """The worksheet, or None when the feature is not configured.

    Only successful connections stay cached: Streamlit Cloud applies new
    secrets WITHOUT restarting the app, so a None cached at boot (before the
    secrets existed) would otherwise stick until a manual reboot.
    """
    ws = _cached_ws()
    if ws is None:
        _cached_ws.clear()
    return ws


def versions_available() -> bool:
    return _versions_ws() is not None


@st.cache_data(ttl=45, show_spinner=False)
def fetch_versions(phone: str) -> list[dict]:
    """All saved versions for a phone, newest first."""
    ws = _versions_ws()
    if ws is None:
        return []
    rows = [r for r in ws.get_all_records() if str(r.get("phone", "")) == phone]
    return list(reversed(rows))


def result_summary(out: dict) -> str:
    if out["kind"] == "success":
        if out.get("target_age") is not None:
            return (f"{fmt_mon_yyyy(out['retirement_date'])} "
                    f"(age {out['target_age']:.1f}): feasible")
        return (f"Retire {fmt_mon_yyyy(out['retirement_date'])} "
                f"(age {out['age_at_retirement']:.1f})")
    if out["kind"] == "target_infeasible":
        sip = out.get("sip_needed")
        if sip:
            return (f"{fmt_mon_yyyy(out['chosen_date'])}: "
                    f"needs +{format_inr(sip)}/mo")
        return f"{fmt_mon_yyyy(out['chosen_date'])}: not achievable via SIP"
    if out["kind"] == "infeasible":
        f = out.get("failure") or {}
        when = fmt_mon_yyyy(f["date"]) if f.get("date") else "?"
        return f"Infeasible — fails {when}"
    return "Validation error"


def save_version(phone: str, note: str, config: dict, out: dict) -> str | None:
    """Append one row for this run; returns the version label, or None."""
    ws = _versions_ws()
    if ws is None:
        return None
    try:
        version = f"v{len(fetch_versions(phone)) + 1}"
        saved_at = pd.Timestamp.now(tz="Asia/Kolkata").strftime("%Y-%m-%d %H:%M")
        ws.append_row(
            [phone, config.get("client_name", ""), saved_at, version,
             (note or "").strip(), result_summary(out),
             build_inputs_json(config).decode("utf-8"),
             config.get("simulation_id", ""),
             build_output_json(config, out).decode("utf-8")],
            value_input_option="RAW",
        )
        fetch_versions.clear()
        return version
    except Exception as e:
        st.session_state.versions_error = str(e)
        return None


def _apply_sip_to_form(amount: int):
    """st.button callback: turn the solver's answer into a real, visible stream."""
    today = st.session_state.today
    st.session_state.streams.append({
        "_uid": _next_uid(), "name": "Additional SIP", "amount": int(amount),
        "start_date": today, "end_date_mode": "At retirement",
        "end_date": add_years(today, 30), "step_up_percent": 0.0,
        "step_up_frequency": "Annual", "step_up_date": today,
    })
    st.session_state.run_output = None
    st.session_state.upload_msg = (
        "success",
        f"Added an 'Additional SIP' stream of {format_inr(amount)}/month — "
        "press Run simulation to verify the full plan.",
    )


def _apply_selected_version():
    """st.button callback: load the chosen saved version into the form."""
    rows = st.session_state.get("_version_rows") or []
    label = st.session_state.get("version_choice")
    row = next((r for r in rows if r["_label"] == label), None)
    if row is None:
        st.session_state.upload_msg = ("error", "Pick a version first.")
        return
    try:
        doc = json.loads(row["inputs_json"])
        _load_doc_into_form(doc, row["_label"])
    except (ValueError, KeyError) as e:
        st.session_state.upload_msg = ("error", f"Could not load that version: {e}")


# ── Widgets ─────────────────────────────────────────────────────────────────
def month_year_input(container, label: str, ts, key: str) -> pd.Timestamp:
    """Month + year picker (the engine's grid is monthly — day is always the 1st).

    When the widget keys already exist in session_state (seeded at init or set
    programmatically on load), no defaults are passed — the state alone drives
    the widget. Passing a default alongside API-set state is the pattern that
    let real browsers resurrect stale values after a version load.
    """
    ts = pd.Timestamp(ts) if ts is not None else month_start_today()
    c1, c2 = container.columns([3, 2])
    mkw = {} if f"{key}_m" in st.session_state else {"index": ts.month - 1}
    ykw = {} if f"{key}_y" in st.session_state else {"value": int(ts.year)}
    month = c1.selectbox(f"{label} — month", MONTH_NAMES, key=f"{key}_m", **mkw)
    year = c2.number_input(f"{label} — year", min_value=1950, max_value=2150,
                           step=1, key=f"{key}_y", **ykw)
    return pd.Timestamp(int(year), MONTH_NAMES.index(month) + 1, 1)


def money_input(container, label: str, value, key: str, help: str | None = None) -> int:
    vkw = {} if key in st.session_state else {"value": int(value)}
    amt = container.number_input(label, min_value=0, step=50_000,
                                 key=key, help=help, **vkw)
    container.caption(inr_hint(amt))
    return amt


def _next_uid() -> int:
    st.session_state.uid_counter += 1
    return st.session_state.uid_counter


def init_state() -> None:
    if "streams" in st.session_state:
        return
    today = month_start_today()
    st.session_state.uid_counter = 0
    st.session_state.today = today
    st.session_state.streams = [make_default_stream(0, today)]
    st.session_state.goals = [make_goal_from_template("retirement_income", 0, today)]
    st.session_state.one_time = []
    for item in st.session_state.streams + st.session_state.goals:
        item["_uid"] = _next_uid()
    st.session_state.run_output = None
    st.session_state.personal_defaults = {
        "client_name": "",
        "current_date": today,
        "current_age": 30,
        "target_lifetime": 90,
        "current_corpus": 10_000_000,
        "risk_profile": "Balanced",
        "retirement_mode": "earliest",
        "target_date": add_years(today, 25),
        "phone": "",
    }
    # Personal widgets are STATE-DRIVEN: keys seeded here, set directly on
    # version/file load, and the widgets carry no value=/index= defaults.
    # (The old pop-and-default pattern let real browsers resurrect stale
    # widget values after a load.)
    st.session_state.setdefault("p_client", "")
    st.session_state.setdefault("p_curdate_m", MONTH_NAMES[today.month - 1])
    st.session_state.setdefault("p_curdate_y", int(today.year))
    st.session_state.setdefault("p_age", 30)
    st.session_state.setdefault("p_life", 90)
    st.session_state.setdefault("p_corpus", 10_000_000)
    st.session_state.setdefault("p_risk", "Balanced")
    st.session_state.setdefault("p_mode", "Earliest possible")
    _tgt = add_years(today, 25)
    st.session_state.setdefault("p_target_m", MONTH_NAMES[_tgt.month - 1])
    st.session_state.setdefault("p_target_y", int(_tgt.year))
    st.session_state.setdefault("p_phone", "")


# ── Form sections ───────────────────────────────────────────────────────────
def render_stream(s: dict) -> None:
    uid = s["_uid"]
    r1c1, r1c2 = st.columns([2, 2])
    s["name"] = r1c1.text_input("Name", value=s["name"], key=f"st_name_{uid}")
    s["amount"] = money_input(r1c2, "Monthly amount (₹, as of start date)", s["amount"], f"st_amt_{uid}")
    s["start_date"] = month_year_input(st, "Starts", s["start_date"], f"st_start_{uid}")
    r2c1, r2c2 = st.columns([2, 2])
    s["end_date_mode"] = r2c1.selectbox(
        "Ends", INVESTMENT_END_MODES,
        index=INVESTMENT_END_MODES.index(s["end_date_mode"]), key=f"st_endmode_{uid}",
    )
    if s["end_date_mode"] == "Fixed":
        s["end_date"] = month_year_input(r2c2, "End", s["end_date"], f"st_end_{uid}")
    r3c1, r3c2 = st.columns([2, 2])
    s["step_up_percent"] = r3c1.number_input(
        "Step-up %", min_value=0.0, max_value=100.0, value=float(s["step_up_percent"]),
        step=0.5, key=f"st_supct_{uid}",
    )
    s["step_up_frequency"] = r3c2.selectbox(
        "Step-up frequency", STEPUP_FREQUENCIES,
        index=STEPUP_FREQUENCIES.index(s["step_up_frequency"]), key=f"st_sufreq_{uid}",
    )
    s["step_up_date"] = month_year_input(st, "Step-up anchor", s["step_up_date"], f"st_sudate_{uid}")


def render_goal(g: dict) -> None:
    uid = g["_uid"]
    r1c1, r1c2 = st.columns([2, 2])
    g["name"] = r1c1.text_input("Name", value=g["name"], key=f"g_name_{uid}")
    g["description"] = r1c2.text_input("Description", value=g["description"], key=f"g_desc_{uid}")

    # v2 (2026-08-24): Nature (Replenishing / Non-replenishing) is NOT an input
    # any more — it is derived from whether the goal has cashflows beyond the
    # grid's reach. The two questions that remain are Structure (one-time or
    # recurring — it decides the cashflow series) and Negotiability, shown as
    # "Type" (it picks the grid column). See "4. Goal Planning" §4.4.
    r2c1, r2c2 = st.columns(2)
    g["structure"] = r2c1.selectbox(
        "Structure", GOAL_STRUCTURES,
        index=GOAL_STRUCTURES.index(g["structure"] if g["structure"] in GOAL_STRUCTURES else "Lumpsum"),
        key=f"g_struct_{uid}",
        help="Lumpsum = one payment. Recurring = a series — fees, an income "
             "stream, anything repeating. A goal with both a lump and a stream "
             "is entered as two goals.",
    )
    g["type"] = r2c2.selectbox(
        "Type", GOAL_TYPES, index=GOAL_TYPES.index(g["type"]), key=f"g_type_{uid}",
        help="Negotiability. It sets how early the goal is pre-funded and how "
             "much sits in debt vs hybrid: non-negotiable starts 5 years out, "
             "semi-negotiable 4, negotiable 3. It also sets funding priority "
             "when goals compete for the same money.",
    )

    r3c1, r3c2 = st.columns([2, 2])
    g["start_date_mode"] = r3c1.selectbox(
        "Start", GOAL_START_MODES, index=GOAL_START_MODES.index(g["start_date_mode"]),
        key=f"g_startmode_{uid}",
    )
    if g["start_date_mode"] == "Fixed":
        g["start_date"] = month_year_input(r3c2, "Start date", g["start_date"], f"g_start_{uid}")

    r4c1, r4c2 = st.columns([2, 2])
    g["amount"] = money_input(
        r4c1, "Amount (today's ₹)", g["amount"], f"g_amt_{uid}",
        help="Present value — the engine grows it to the goal date at the growth % below.",
    )
    g["inflation_percent"] = r4c2.number_input(
        "Annual growth %", min_value=0.0, max_value=100.0,
        value=float(g["inflation_percent"]), step=0.5, key=f"g_infl_{uid}",
    )

    _cats = CRM_GOAL_CATEGORIES
    _cur = g.get("goal_category") if g.get("goal_category") in _cats else None
    g["goal_category"] = st.selectbox(
        "Category (for the CRM)", _cats, index=_cats.index(_cur),
        format_func=lambda v: "— not set —" if v is None else v.title(),
        key=f"g_cat_{uid}",
        help="The CRM's own goal taxonomy. Optional, but it is what their "
             "reporting groups goals by.",
    )
    if g.get("purpose_id"):
        st.caption(f"CRM id: `{g['purpose_id']}`")

    if g["structure"] == "Recurring":
        r5c1, r5c2, r5c3 = st.columns(3)
        g["frequency"] = r5c1.selectbox(
            "Frequency", RECURRING_FREQUENCIES,
            index=RECURRING_FREQUENCIES.index(g["frequency"] if g["frequency"] in RECURRING_FREQUENCIES else "Monthly"),
            key=f"g_freq_{uid}",
        )
        g["end_mode"] = r5c2.selectbox(
            "End mode", GOAL_END_MODES,
            index=GOAL_END_MODES.index(g["end_mode"] if g["end_mode"] in GOAL_END_MODES else "Occurrences"),
            key=f"g_endmode_{uid}",
        )
        # Fixed-vs-inflating is POLICY, derived from the goal's shape - the
        # CM sees which rule applies but does not choose it.
        if payments_fixed_for(g):
            st.caption(
                ":lock: **Payments are fixed once the goal starts** - the "
                "amount grows at the growth % only until the first payment, "
                "then every payment stays at that amount (fees lock at "
                "admission, EMIs are signed)."
            )
        else:
            st.caption(
                ":chart_with_upwards_trend: **Income goal - payments keep "
                "growing** at the growth % throughout, because income must "
                "track the cost of living. (A goal counts as income when it "
                "starts at retirement or runs for Lifetime.)"
            )
        if g["end_mode"] == "Occurrences":
            g["occurrences"] = r5c3.number_input(
                "Number of payments", min_value=1, value=int(g["occurrences"] or 1),
                step=1, key=f"g_occ_{uid}",
            )
        elif g["end_mode"] == "Fixed date":
            g["end_date"] = month_year_input(
                r5c3, "End date", g["end_date"] or g["start_date"], f"g_end_{uid}"
            )


def render_one_time(w: dict) -> None:
    uid = w["_uid"]
    c1, c2 = st.columns([2, 2])
    w["name"] = c1.text_input("Name", value=w["name"], key=f"ot_name_{uid}")
    w["amount"] = money_input(c2, "Amount (₹ on that date)", w["amount"], f"ot_amt_{uid}")
    w["date"] = month_year_input(st, "Date", w["date"], f"ot_date_{uid}")


def build_config(personal: dict) -> dict:
    """Assemble the engine's plain-dict config (mirror planForm.buildConfig)."""
    streams = []
    for s in st.session_state.streams:
        streams.append({
            "name": s["name"],
            "amount": float(s["amount"]),
            "start_date": s["start_date"],
            "end_date_mode": s["end_date_mode"],
            "end_date": s["end_date"] if s["end_date_mode"] == "Fixed" else None,
            "step_up_percent": float(s["step_up_percent"]),
            "step_up_frequency": s["step_up_frequency"],
            "step_up_date": s["step_up_date"],
        })
    goals = []
    for raw in st.session_state.goals:
        g = normalise_goal(raw)
        goals.append({
            "name": g["name"],
            "description": g["description"] or "",
            "type": g["type"],
            # Derived, not asked (§4.4) — kept in the config so saved runs and
            # the CRM export still carry the field.
            "nature": ("Replenishing"
                       if goal_purpose(g, personal["current_date"]).startswith("Extends")
                       else "Non-replenishing"),
            "structure": g["structure"],
            "start_date_mode": g["start_date_mode"],
            "start_date": None if g["start_date_mode"] == "At retirement" else g["start_date"],
            "amount": float(g["amount"]),
            "frequency": g["frequency"],
            "occurrences": int(g["occurrences"]) if g.get("occurrences") is not None else None,
            "end_mode": g["end_mode"],
            "end_date": g["end_date"],
            "inflation_percent": float(g["inflation_percent"]),
            "payments_fixed_at_start": payments_fixed_for(g),
            "purpose_id": g.get("purpose_id") or None,
            "goal_category": g.get("goal_category") or None,
        })
    one_time = [
        {"name": w["name"], "date": w["date"], "amount": float(w["amount"])}
        for w in st.session_state.one_time
    ]
    return {
        "current_date": personal["current_date"],
        "current_age": float(personal["current_age"]),
        "target_lifetime": float(personal["target_lifetime"]),
        "current_corpus": float(personal["current_corpus"]),
        "risk_profile": personal["risk_profile"],
        # Read only by the Excel export headers — the engine ignores these.
        "client_name": personal["client_name"] or "Playground",
        "m3_id": "playground",
        "phone": personal.get("phone") or "",
        "retirement_mode": personal.get("retirement_mode") or "earliest",
        "target_date": personal.get("target_date"),
        "investment_streams": streams,
        "goals": goals,
        "one_time_investments": one_time,
    }


def build_output_json(config: dict, out: dict) -> bytes:
    """Compact result record for the simulation log (pairs with the inputs JSON).

    Deliberately headline-level — verdict, dates, snapshot, key wealth figures,
    failure/validation detail — not the monthly ledger (which lives in the
    CSV/Excel and would not fit a sheet cell).
    """
    doc = {
        "simulation_id": config.get("simulation_id", ""),
        "generated_at": pd.Timestamp.now(tz="Asia/Kolkata").strftime("%Y-%m-%d %H:%M:%S"),
        "engine_version": ENGINE_SOURCE_SHA,
        "glidepath_version": GLIDEPATH_VERSION,
        "client_name": config.get("client_name", ""),
        "phone": config.get("phone", ""),
        "risk_profile": config.get("risk_profile", ""),
        "retirement_mode": config.get("retirement_mode", "earliest"),
        "verdict": out["kind"],
    }
    if out.get("target_age") is not None:
        doc["target_age"] = float(out["target_age"])
    if out.get("earliest_date") is not None:
        doc["earliest_feasible_date"] = _iso_or_none(out["earliest_date"])
    if out["kind"] == "success":
        wealth = out.get("wealth")
        doc.update({
            "retirement_date": _iso_or_none(out["retirement_date"]),
            "age_at_retirement": round(out["age_at_retirement"], 1),
            "snapshot_at_retirement": out.get("snapshot"),
        })
        if wealth is not None and not wealth.empty:
            total = wealth["Total wealth"]
            doc["wealth_at_plan_end"] = round(float(total.iloc[-1]), 2)
            doc["peak_wealth"] = {
                "amount": round(float(total.max()), 2),
                "date": total.idxmax().strftime("%Y-%m-01"),
            }
    elif out["kind"] == "target_infeasible":
        f = out.get("failure") or {}
        doc.update({
            "target_date": _iso_or_none(out["chosen_date"]),
            "sip_needed_monthly": out.get("sip_needed"),
            "monthly_investment_now": round(out.get("monthly_now", 0), 2),
            "failure": {
                "date": _iso_or_none(f.get("date")),
                "description": str(f.get("description", "Corpus depletion")),
            },
        })
    elif out["kind"] == "infeasible":
        f = out.get("failure") or {}
        doc["failure"] = {
            "date": _iso_or_none(f.get("date")),
            "description": str(f.get("description", "Corpus depletion")),
        }
    else:
        doc["validation_errors"] = list(out.get("errors") or [])
    return json.dumps(doc, indent=2, ensure_ascii=False).encode("utf-8")


# ── Target-age mode + SIP-needed search ─────────────────────────────────────
# DECISIONS.md (2026-06-02) removed v3's target-date mode from the solver and
# prescribed exactly this reintroduction: a thin wrapper that calls
# run_simulation() once for the chosen date, keeping find_retirement_date()
# solver-only. The engine stays untouched. Direct run_simulation callers must
# validate the config themselves — the solver did that for us until now.

SIP_SEARCH_START = 25_000       # first doubling probe (₹/month)
SIP_SEARCH_CAP = 10_000_000     # ₹1 Cr/month — beyond this, report "not fixable by SIP"
SIP_ROUND_TO = 500              # present a clean, verified number


def target_retirement_date(config: dict) -> pd.Timestamp:
    """The chosen retirement month, snapped to the month grid."""
    t = pd.Timestamp(config["target_date"])
    return pd.Timestamp(t.year, t.month, 1)


def target_age_of(config: dict, chosen: pd.Timestamp) -> float:
    """Derived age at the chosen date (input is a month/year since 2026-08-26)."""
    current = pd.Timestamp(config["current_date"])
    return float(config["current_age"]) + (chosen - current).days / 365.25


def _extra_sip_stream(config: dict, amount: float) -> dict:
    """The solver's synthetic stream: flat (no step-up), today → retirement.

    Never written to any export — when the CM accepts the number, the Apply
    button adds a real, visible stream instead.
    """
    today = pd.Timestamp(config["current_date"])
    return {
        "name": "Additional SIP (solver)", "amount": float(amount),
        "start_date": today, "end_date_mode": "At retirement", "end_date": None,
        "step_up_percent": 0.0, "step_up_frequency": "Annual", "step_up_date": today,
    }


def sip_needed_search(config: dict, chosen_date: pd.Timestamp, instrument_params,
                      glide_paths, progress=None) -> int | None:
    """Minimum additional flat monthly SIP that makes chosen_date feasible.

    Feasibility is monotone in money-in, so: double an upper bound until
    feasible (capped), binary-search to ₹500, round UP, verify the final
    number with one more full simulation. Returns None when even the cap
    cannot rescue the plan.
    """
    def feasible(amount: float) -> bool:
        if progress:
            progress(f"Testing +{format_inr(amount)}/month …")
        trial = dict(config)
        trial["investment_streams"] = list(config.get("investment_streams") or []) \
            + [_extra_sip_stream(config, amount)]
        ok, *_rest = run_simulation(trial, chosen_date, instrument_params, glide_paths)
        return bool(ok)

    lo, hi = 0.0, float(SIP_SEARCH_START)
    while not feasible(hi):
        lo, hi = hi, hi * 2
        if hi > SIP_SEARCH_CAP:
            return None
    while hi - lo > SIP_ROUND_TO:
        mid = (lo + hi) / 2
        if feasible(mid):
            hi = mid
        else:
            lo = mid
    import math
    amount = int(math.ceil(hi / SIP_ROUND_TO) * SIP_ROUND_TO)
    while not feasible(amount):            # guard the rounding edge, cheaply
        amount += SIP_ROUND_TO
    return amount


def monthly_investment_today(config: dict) -> float:
    current = pd.Timestamp(config["current_date"])
    return sum(float(s.get("amount", 0) or 0)
               for s in config.get("investment_streams") or []
               if pd.Timestamp(s["start_date"]) <= current)


def run_plan_target(config: dict, progress=None) -> dict:
    """Chosen-age mode: feasibility at that date; if infeasible, the SIP gap."""
    instrument_params = resolve_instrument_params(config["risk_profile"])
    glide_paths = get_glide_paths()
    current_date = pd.Timestamp(config["current_date"])
    death_date = current_date + pd.DateOffset(
        years=int(config["target_lifetime"] - config["current_age"])
    )
    try:
        validate_plan_config(config)
    except PlanValidationError as e:
        return {"kind": "invalid", "errors": list(e.errors), "config": config}

    chosen = target_retirement_date(config)
    if not (current_date < chosen < death_date):
        return {"kind": "invalid", "config": config, "errors": [
            f"Chosen retirement month {fmt_mon_yyyy(chosen)} must be after the "
            f"plan start ({fmt_mon_yyyy(current_date)}) and before the plan end "
            f"({fmt_mon_yyyy(death_date)})."]}
    target_age = target_age_of(config, chosen)

    # Context both branches want: the earliest date the CURRENT plan supports.
    solved = find_retirement_date(config, instrument_params, glide_paths)
    earliest = pd.Timestamp(solved["retirement_date"]) if solved["success"] else None

    success, _t, failure, pools_df, goal_dfs, comp_df = run_simulation(
        config, chosen, instrument_params, glide_paths
    )

    if success:
        snapshot = build_snapshot(comp_df, chosen) if not comp_df.empty else None
        workbook = build_advisor_workbook(
            config, {"success": True, "retirement_date": chosen, "failure": None},
            comprehensive_df=comp_df, snapshot=snapshot,
            goal_dfs=goal_dfs, pool_movements_df=pools_df,
        )
        workbook = add_summary_sheet(
            workbook, config, kind="success", retirement_date=chosen,
            age_at_retirement=target_age, comp_df=comp_df, snapshot=snapshot,
            death_date=death_date, mode_info={"target_age": target_age,
                                              "earliest": earliest},
        )
        return {
            "kind": "success", "config": config,
            "retirement_date": chosen, "age_at_retirement": target_age,
            "target_age": target_age, "earliest_date": earliest,
            "snapshot": snapshot, "wealth": wealth_frame(comp_df, death_date),
            "goal_table": build_goal_results(config, chosen),
            "workbook": workbook, "csv": csv_with_summary(comp_df),
        }

    sip = sip_needed_search(config, chosen, instrument_params, glide_paths, progress)
    # retirement_date = the tested date: the diagnostic ran exactly there, and
    # it lets the export resolve at-retirement goals (None would NaT-crash it).
    diag_result = {"success": False, "retirement_date": chosen, "failure": failure}
    try:
        workbook = build_advisor_workbook(
            config, diag_result, comprehensive_df=comp_df, snapshot=None,
            goal_dfs=goal_dfs, pool_movements_df=pools_df,
        )
        workbook = add_summary_sheet(
            workbook, config, kind="target_infeasible", failure=failure,
            comp_df=comp_df, death_date=death_date,
            mode_info={"target_age": target_age, "chosen": chosen,
                       "earliest": earliest, "sip_needed": sip,
                       "monthly_now": monthly_investment_today(config)},
        )
    except Exception:
        workbook = None
    return {
        "kind": "target_infeasible", "config": config,
        "target_age": target_age, "chosen_date": chosen,
        "earliest_date": earliest, "sip_needed": sip,
        "monthly_now": monthly_investment_today(config),
        "failure": failure,
        "goal_table": build_goal_results(config, None),
        "workbook": workbook,
        "csv": csv_with_summary(comp_df)
        if comp_df is not None and not comp_df.empty else None,
    }


def run_plan(config: dict, progress=None) -> dict:
    """Solve + simulate; returns everything the results pane needs."""
    config["simulation_id"] = new_simulation_id()
    if config.get("retirement_mode") == "target_age":
        return run_plan_target(config, progress=progress)
    instrument_params = resolve_instrument_params(config["risk_profile"])
    glide_paths = get_glide_paths()
    current_date = pd.Timestamp(config["current_date"])
    death_date = current_date + pd.DateOffset(
        years=int(config["target_lifetime"] - config["current_age"])
    )

    try:
        solved = find_retirement_date(config, instrument_params, glide_paths)
    except PlanValidationError as e:
        return {"kind": "invalid", "errors": list(e.errors)}

    if not solved["success"]:
        # Infeasible: diagnostic run at the lifetime end (mirrors service.py).
        # Its outputs are still worth downloading — the workbook carries the
        # failure rows, and the comprehensive view exists for the corpus-
        # depletion failure class (it is empty by design for debt-pool
        # depletion, in which case the CSV is simply not offered).
        _s, _t, failure, pools_df, goal_dfs, comp_df = run_simulation(
            config, death_date, instrument_params, glide_paths
        )
        # The diagnostic ran AT death_date — pass it so the export can resolve
        # at-retirement goals (retirement_date=None NaT-crashed _goals_sheet,
        # which used to degrade these workbooks to None).
        diag_result = {"success": False, "retirement_date": death_date,
                       "failure": failure}
        try:
            workbook = build_advisor_workbook(
                config, diag_result, comprehensive_df=comp_df, snapshot=None,
                goal_dfs=goal_dfs, pool_movements_df=pools_df,
            )
            workbook = add_summary_sheet(
                workbook, config, kind="infeasible", failure=failure,
                comp_df=comp_df, death_date=death_date,
            )
        except Exception:
            workbook = None  # degrade to no workbook rather than a dead screen
        return {
            "kind": "infeasible",
            "config": config,
            "failure": failure,
            "solver_failure": solved.get("failure"),
            "goal_table": build_goal_results(config, None),
            "workbook": workbook,
            "csv": csv_with_summary(comp_df)
            if comp_df is not None and not comp_df.empty else None,
        }

    retirement_date = pd.Timestamp(solved["retirement_date"])
    success, _trans, failure, pools_df, goal_dfs, comp_df = run_simulation(
        config, retirement_date, instrument_params, glide_paths
    )
    snapshot = build_snapshot(comp_df, retirement_date) if not comp_df.empty else None
    wealth = wealth_frame(comp_df, death_date)
    age_at_ret = float(config["current_age"]) + (retirement_date - current_date).days / 365.25

    workbook = build_advisor_workbook(
        config, solved, comprehensive_df=comp_df, snapshot=snapshot,
        goal_dfs=goal_dfs, pool_movements_df=pools_df,
    )
    workbook = add_summary_sheet(
        workbook, config, kind="success", retirement_date=retirement_date,
        age_at_retirement=age_at_ret, comp_df=comp_df, snapshot=snapshot,
        death_date=death_date,
    )
    return {
        "kind": "success",
        "config": config,
        "retirement_date": retirement_date,
        "age_at_retirement": age_at_ret,
        "snapshot": snapshot,
        "wealth": wealth,
        "goal_table": build_goal_results(config, retirement_date),
        "workbook": workbook,
        "csv": csv_with_summary(comp_df),
    }


# ── Results pane ────────────────────────────────────────────────────────────
def render_results(out: dict) -> None:
    if out["kind"] == "invalid":
        st.error("The plan inputs failed validation:\n\n" +
                 "\n".join(f"- {e}" for e in out["errors"]))
        return

    if out["kind"] == "infeasible":
        st.error("This plan is **not fundable** within the target lifetime — even "
                 "retiring at the very end of the plan horizon, the corpus fails.")
        failure = out.get("failure") or {}
        if failure:
            st.warning(
                f"First failure: **{fmt_mon_yyyy(failure.get('date'))}** — "
                f"{failure.get('description', 'Corpus depletion')}"
            )
        if not out["goal_table"].empty:
            st.dataframe(out["goal_table"], use_container_width=True, hide_index=True)
        st.caption("Try: lower goal amounts, later goal dates, higher investments, "
                   "or a more aggressive risk profile.")
        st.subheader("Downloads")
        d1, d2, d3 = st.columns(3)
        if out.get("workbook"):
            d1.download_button(
                "📗 Advisor workbook (Excel)", data=out["workbook"],
                file_name="financial_plan_advisor.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl_wb_infeasible",
            )
        if out.get("csv"):
            d2.download_button(
                "📄 Comprehensive monthly (CSV)", data=out["csv"],
                file_name="financial_plan_monthly.csv", mime="text/csv",
                key="dl_csv_infeasible",
            )
        d3.download_button(
            "📥 Inputs (JSON)", data=build_inputs_json(out["config"]),
            file_name=inputs_filename(out["config"]), mime="application/json",
            key="dl_inputs_infeasible",
        )
        st.caption(
            "Diagnostic outputs: no feasible retirement exists, so the workbook/CSV "
            "reflect a run at the latest possible retirement (lifetime end), "
            "including the failure and shortfall rows."
        )
        return

    if out["kind"] == "target_infeasible":
        age = out["target_age"]
        chosen = out["chosen_date"]
        sip = out.get("sip_needed")
        st.error(f"Retiring in **{fmt_mon_yyyy(chosen)}** (age {age:.1f}) is "
                 "**not fundable** with the current plan.")
        if sip:
            now = out.get("monthly_now", 0)
            m1, m2, m3 = st.columns(3)
            m1.metric("Additional SIP needed", f"{short_inr(sip)}/mo")
            m2.metric("Monthly investment", f"{short_inr(now)} → {short_inr(now + sip)}")
            m3.metric("Current plan supports",
                      fmt_mon_yyyy(out["earliest_date"])
                      if out.get("earliest_date") is not None else "no age")
            st.caption(
                f"A flat additional SIP of **{format_inr(sip)}/month** (no step-up, "
                f"starting now, until retirement) makes {fmt_mon_yyyy(chosen)} "
                "feasible — verified by a full simulation at exactly that amount."
            )
            st.button(
                f"➕ Apply +{format_inr(sip)}/month to the form",
                on_click=_apply_sip_to_form, args=(sip,),
            )
        else:
            st.warning(
                "No realistic additional SIP fixes this date — the shortfall is "
                "structural (goals due too soon or too large). Revisit goal "
                "amounts/dates, or pick a later retirement age."
                + (f" The current plan supports "
                   f"**{fmt_mon_yyyy(out['earliest_date'])}**."
                   if out.get("earliest_date") is not None else "")
            )
        failure = out.get("failure") or {}
        if failure:
            st.caption(f"Without the extra SIP, first failure: "
                       f"{fmt_mon_yyyy(failure.get('date'))} — "
                       f"{failure.get('description', 'Corpus depletion')}")
        if not out["goal_table"].empty:
            st.dataframe(out["goal_table"], use_container_width=True, hide_index=True)
        st.subheader("Downloads")
        d1, d2, d3 = st.columns(3)
        if out.get("workbook"):
            d1.download_button(
                "📗 Advisor workbook (Excel)", data=out["workbook"],
                file_name="financial_plan_advisor.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl_wb_target",
            )
        if out.get("csv"):
            d2.download_button(
                "📄 Comprehensive monthly (CSV)", data=out["csv"],
                file_name="financial_plan_monthly.csv", mime="text/csv",
                key="dl_csv_target",
            )
        d3.download_button(
            "📥 Inputs (JSON)", data=build_inputs_json(out["config"]),
            file_name=inputs_filename(out["config"]), mime="application/json",
            key="dl_inputs_target",
        )
        st.caption("Diagnostic outputs at the chosen retirement date, without the "
                   "extra SIP. Use 'Apply to the form' and re-run for the full "
                   "feasible-plan artifacts.")
        return

    ret = out["retirement_date"]
    snap = out["snapshot"] or {}
    if out.get("target_age") is not None:
        extra = (f" · earliest possible: **{fmt_mon_yyyy(out['earliest_date'])}**"
                 if out.get("earliest_date") is not None else "")
        st.success(f"Retiring in **{fmt_mon_yyyy(ret)}** "
                   f"(age {out['target_age']:.1f}) is **feasible**{extra}")
    else:
        st.success(f"Earliest feasible retirement: **{fmt_mon_yyyy(ret)}** "
                   f"(age {out['age_at_retirement']:.1f})")

    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Retirement", fmt_mon_yyyy(ret))
    m2.metric("Age at retirement", f"{out['age_at_retirement']:.1f}")
    m3.metric("Wealth at retirement", short_inr(snap.get("total", 0)))
    if not out["wealth"].empty:
        m4.metric("Wealth at lifetime end", short_inr(float(out["wealth"]["Total wealth"].iloc[-1])))

    st.line_chart(out["wealth"], use_container_width=True)

    with st.expander("Wealth snapshot at retirement"):
        if snap:
            rows = [
                ("Core corpus", snap["core"]), ("Debt pool", snap["debt"]),
                ("Hybrid pool", snap["hybrid"]), ("Goal debt tranches", snap["goal_debt"]),
                ("Goal hybrid tranches", snap["goal_hybrid"]), ("Total", snap["total"]),
            ]
            st.table(pd.DataFrame(
                [(k, format_inr(v)) for k, v in rows], columns=["Bucket", "Value"]
            ))

    st.subheader("Goals")
    st.dataframe(out["goal_table"], use_container_width=True, hide_index=True)

    st.subheader("Downloads")
    d1, d2, d3, d4 = st.columns(4)
    d4.download_button(
        "🔗 CRM goals upload (JSON)",
        data=crm_goals_upload_json(out["config"], out["retirement_date"]),
        file_name=crm_upload_filename(out["config"]), mime="application/json",
        key="dl_crm_upload",
        help="The CRM's strict goals file: flat rows, resolved dates, their "
             "vocabulary. Upload this straight into the CRM.",
    )
    d1.download_button(
        "📗 Advisor workbook (Excel)", data=out["workbook"],
        file_name="financial_plan_advisor.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    d2.download_button(
        "📄 Comprehensive monthly (CSV)", data=out["csv"],
        file_name="financial_plan_monthly.csv", mime="text/csv",
    )
    d3.download_button(
        "📥 Inputs (JSON)", data=build_inputs_json(out["config"], out["retirement_date"]),
        file_name=inputs_filename(out["config"]), mime="application/json",
        key="dl_inputs_success",
    )
    st.caption(
        "**Inputs (JSON)** is our own record — it reloads into this form. "
        "**CRM goals upload** is the CRM's strict contract: flat rows in their "
        "vocabulary, 'At retirement' resolved to the solved date, and a "
        f"lifetime series written as {CRM_OPEN_ENDED_OCCURRENCES} payments. "
        "Load a CRM goals file back here to pick up the ids they mint."
    )


# ── Page ────────────────────────────────────────────────────────────────────
def main() -> None:
    st.set_page_config(page_title="Financial Planning Playground", page_icon="🧮", layout="wide")
    init_state()
    today = st.session_state.today

    st.title("🧮 Financial Planning Playground")
    # NB: this engine is no longer the untouched handoff copy — it is the sole
    # Financial Plan build and has deliberate changes on top (DECISIONS.md).
    st.caption(
        f"Engine `{ENGINE_SOURCE_SHA}` · **updated {ENGINE_UPDATED}** — the "
        "sole Financial Plan build. Goals are provisioned by the **goal grid** "
        "(v2): each goal cashflow is pre-funded across debt and hybrid by how "
        "far away it is and how negotiable the goal is. Every engine change is "
        "logged in `v3_docs/DECISIONS.md`."
    )

    # Sidebar: personal & corpus + risk profile. Defaults come from
    # personal_defaults so a loaded JSON can re-seed every widget.
    d = st.session_state.personal_defaults
    with st.sidebar:
        st.header("Personal & Corpus")
        client_name = st.text_input("Client name (Excel header only)", key="p_client")
        current_date = month_year_input(st, "Plan start", d["current_date"], "p_curdate")
        c1, c2 = st.columns(2)
        current_age = c1.number_input("Current age", min_value=0, max_value=110,
                                      step=1, key="p_age")
        target_lifetime = c2.number_input("Target lifetime", min_value=1, max_value=120,
                                          step=1, key="p_life")
        current_corpus = money_input(st, "Current corpus (₹)", d["current_corpus"], "p_corpus")
        risk_profile = st.selectbox("Risk profile", RISK_PROFILES, key="p_risk")
        st.caption(
            f"Core-corpus return the engine will use: "
            f"**{RISK_PROFILE_CORE_RETURNS[risk_profile] * 100:g}%** · "
            f"debt {_DEFAULT_INSTRUMENT_PARAMS['debt']['return'] * 100:g}%, "
            f"hybrid {_DEFAULT_INSTRUMENT_PARAMS['hybrid']['return'] * 100:g}%."
        )
        # These two assumptions are fundamental to every number the tool
        # produces, so they are stated in the UI rather than buried in code.
        # Values are read from the engine itself so they can never drift.
        with st.expander("Model assumptions - returns & taxation"):
            _dp = _DEFAULT_INSTRUMENT_PARAMS
            st.markdown(
                "**Returns (annual)**\n\n"
                f"- Core corpus (equity): **{RISK_PROFILE_CORE_RETURNS[risk_profile] * 100:g}%**"
                " - set by the risk profile\n"
                f"- Hybrid: **{_dp['hybrid']['return'] * 100:g}%**\n"
                f"- Debt: **{_dp['debt']['return'] * 100:g}%**\n\n"
                "**Taxation - applied on every redemption, per FIFO tax lot**\n\n"
                f"- **All buckets are equity-taxed**: "
                f"{_dp['hybrid']['stcg_tax'] * 100:g}% on gains held under a year, "
                f"{_dp['hybrid']['ltcg_tax'] * 100:g}% beyond. The debt bucket "
                "holds **arbitrage funds** (debt-like return, equity taxation) "
                "and the hybrid funds offered are equity-taxed.\n"
                "- **Year-boundary rule**: a redemption within 1-2 days of "
                f"completing a year is taxed as long-term "
                f"({_dp['hybrid']['ltcg_tax'] * 100:g}%) - the desk shifts the "
                "redemption to cross the year.\n\n"
                "Tax is charged whenever money *moves*, not only when a goal is "
                "paid - including transfers between buckets."
            )
        st.divider()
        st.header("Retirement")
        # Internal mode value stays "target_age" for saved-file compatibility;
        # the input is a month/year since 2026-08-26 (age shows as a caption).
        _mode_labels = {"earliest": "Earliest possible", "target_age": "At a chosen date"}
        mode_label = st.radio(
            "Retirement date", list(_mode_labels.values()),
            key="p_mode", horizontal=True,
            help="Earliest possible = the solver finds the first feasible month. "
                 "At a chosen date = test that month; if it is not fundable, the "
                 "app computes the additional SIP needed.",
        )
        retirement_mode = "earliest" if mode_label == _mode_labels["earliest"] else "target_age"
        target_date = None
        if retirement_mode == "target_age":
            target_date = month_year_input(
                st, "Retire in", d.get("target_date") or add_years(current_date, 25),
                "p_target",
            )
            _yrs = (target_date - current_date).days / 365.25
            if target_date <= current_date:
                st.caption(":red[Pick a month after the plan start.]")
            elif current_age + _yrs >= target_lifetime:
                st.caption(":red[That is at or beyond the target lifetime.]")
            else:
                st.caption(f"= age **{current_age + _yrs:.1f}**")
        st.divider()
        st.header("Client versions")
        phone_raw = st.text_input("Client phone number", key="p_phone",
                                  placeholder="10-digit mobile",
                                  help="Keys the saved-version history. Every Run "
                                       "with a valid number saves a version.")
        phone = normalize_phone(phone_raw)
        note = ""
        if phone_raw and phone is None:
            st.caption(":red[Enter a 10-digit mobile number.]")
        if versions_available():
            if phone:
                note = st.text_input("Note for next save (optional)", key="p_note",
                                     placeholder="e.g. final shown to client")
                rows = fetch_versions(phone)
                for r in rows:
                    bits = [str(r.get("version", "")), str(r.get("saved_at", ""))[5:],
                            str(r.get("result", ""))]
                    if r.get("note"):
                        bits.append(str(r["note"]))
                    r["_label"] = " · ".join(b for b in bits if b)
                st.session_state._version_rows = rows
                if rows:
                    st.selectbox(f"{len(rows)} saved version(s)",
                                 [r["_label"] for r in rows], key="version_choice")
                    st.button("Load selected version", use_container_width=True,
                              on_click=_apply_selected_version)
                else:
                    st.caption("No versions yet for this number — run a "
                               "simulation to save the first one.")
        else:
            err = st.session_state.get("versions_error")
            st.caption("Version history is not configured yet"
                       + (f" (error: {err})" if err else "")
                       + " — runs are not being saved.")
        st.divider()
        with st.expander("📂 Load inputs JSON"):
            st.file_uploader(
                "A financial_plan_inputs_*.json downloaded from this app",
                type=["json"], key="inputs_uploader",
            )
            st.button("Load into form", on_click=_apply_uploaded_inputs,
                      use_container_width=True)
        msg = st.session_state.pop("upload_msg", None)
        if msg is not None:
            (st.success if msg[0] == "success" else st.error)(msg[1])

    personal = {
        "client_name": client_name,
        "current_date": current_date,
        "current_age": current_age,
        "target_lifetime": target_lifetime,
        "current_corpus": current_corpus,
        "risk_profile": risk_profile,
        "phone": phone,
        "retirement_mode": retirement_mode,
        "target_date": target_date,
    }

    left, right = st.columns([1, 1], gap="large")

    with left:
        st.subheader("Investment streams")
        for i, s in enumerate(list(st.session_state.streams)):
            with st.expander(f"💰 {s['name'] or f'Stream {i + 1}'}", expanded=(len(st.session_state.streams) == 1)):
                render_stream(s)
                if st.button("Remove stream", key=f"rm_st_{s['_uid']}"):
                    st.session_state.streams.remove(s)
                    st.rerun()
        if st.button("➕ Add stream"):
            new = make_default_stream(len(st.session_state.streams), today)
            new["_uid"] = _next_uid()
            st.session_state.streams.append(new)
            st.rerun()

        st.subheader("One-time investments")
        for w in list(st.session_state.one_time):
            with st.expander(f"🪙 {w['name'] or 'One-time investment'}", expanded=True):
                render_one_time(w)
                if st.button("Remove", key=f"rm_ot_{w['_uid']}"):
                    st.session_state.one_time.remove(w)
                    st.rerun()
        if st.button("➕ Add one-time investment"):
            st.session_state.one_time.append(
                {"_uid": _next_uid(), "name": f"Investment {len(st.session_state.one_time) + 1}",
                 "date": today, "amount": 500_000}
            )
            st.rerun()

    with right:
        st.subheader("Goals")
        for i, g in enumerate(list(st.session_state.goals)):
            with st.expander(f"🎯 {g['name'] or f'Goal {i + 1}'}", expanded=(len(st.session_state.goals) == 1)):
                render_goal(g)
                if st.button("Remove goal", key=f"rm_g_{g['_uid']}"):
                    st.session_state.goals.remove(g)
                    st.rerun()
        t1, t2 = st.columns([2, 1])
        template_label = t1.selectbox("Goal template", list(GOAL_TEMPLATES.keys()),
                                      label_visibility="collapsed")
        if t2.button("➕ Add goal"):
            new = make_goal_from_template(
                GOAL_TEMPLATES[template_label], len(st.session_state.goals), today
            )
            # Templates always insert the same default name (two children both
            # get "Child Education"), so number the duplicate immediately.
            new["name"] = unique_goal_name(
                new["name"], [g.get("name") for g in st.session_state.goals])
            new["_uid"] = _next_uid()
            st.session_state.goals.append(new)
            st.rerun()
        _dupes = duplicate_goal_names(st.session_state.goals)
        if _dupes:
            st.warning(
                "Two or more goals share a name: **"
                + "**, **".join(_dupes)
                + "**. They are simulated as separate goals and each is fully "
                  "funded — outputs label the repeats “Name #2”, “Name #3”. "
                  "Rename them here if you want your own labels."
            )

    st.divider()
    if st.button("▶ Run simulation", type="primary", use_container_width=True):
        config = build_config(personal)
        spin_text = ("Testing the chosen retirement age…"
                     if retirement_mode == "target_age"
                     else "Solving for the earliest feasible retirement date…")
        progress_slot = st.empty()
        with st.spinner(spin_text):
            st.session_state.run_output = run_plan(
                config, progress=lambda msg: progress_slot.caption(msg)
            )
        progress_slot.empty()
        # Auto-save this run as a version when a valid phone number is set.
        if phone and versions_available():
            version = save_version(phone, note, config, st.session_state.run_output)
            if version:
                st.toast(f"Saved {version} for {phone}", icon="💾")
            else:
                st.warning("This run could not be saved to the version history"
                           + (f": {st.session_state.get('versions_error')}"
                              if st.session_state.get("versions_error") else "."))

    if st.session_state.run_output is not None:
        out = st.session_state.run_output
        sim_id = (out.get("config") or {}).get("simulation_id", "")
        st.caption("Results reflect the inputs at the last Run — re-run after editing."
                   + (f"  ·  {sim_id}" if sim_id else ""))
        render_results(out)


if __name__ == "__main__":
    main()
